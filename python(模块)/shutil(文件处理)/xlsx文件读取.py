import openpyxl
import pandas as pd
import xlrd, xlwt
from xlutils.copy import copy

def openPyxl_way():
    file_dir = r"C:\Users\admin\Desktop\showdata.xlsx"
    fp = openpyxl.load_workbook(file_dir)

    '''获取一张表,默认第一张'''
    # tables = fp.sheetnames
    # print(tables, type(tables))

    '''创建一张表'''
    # fp.create_chartsheet('api_case')
    print(fp.sheetnames)

    '''获取指定的表,'''
    wb = fp['APIcase']
    # print(wb)

    '''对当前选中的表单进行操作。。。。。。。。。。。。。。。。'''

    ws = fp.active
    '''查看第一列的标头'''
    # print(ws['A1'])
    print('序号名称:', ws['A1'].value)

    b = ws['B1']
    print('行是:{}, 列是:{}, 标头是:{}'.format(b.row, b.column, b.value))

    '''获取单元格中所指定的值.....................................'''
    all = ws.cell(row=2, column=2,).value
    print('获取指定的值为:',all)

    for i in range(1, 4):
        '''获取多行的值'''
        print(i, ws.cell(row=i, column=7,).value)

    '''获取指定的行和列,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,,'''
    column = ws['C']    #获取c列的所有数据
    print('c列数据为:', column)

    row2 = ws[2]    #获取第2行的所有数据
    print('第二行数据为:', row2)

    '''遍历单元格中行和列的值______________________________________'''
    row_range = ws[2:3]
    lists = []
    for row_y in row_range:
        for i in row_y:
            lists.append(i.value)
    print('2到3行的数据为:',lists)

    column_range = ws['H:L']
    columns = []
    for f in column_range:
        for j in f:
            columns.append(j.value)
    print('H列到L列的数据为:', columns)

    all_value = []
    for o in ws.iter_rows(min_row=1, max_row=3, min_col=7, max_col=8):
        for l in o:
            all_value.append(l.value)
    print('1-3行,7-8列的数据为：', all_value)

    '''查看行和列的总数'''
    all_row_column = '{}行, {}列'.format(ws.max_row, ws.max_column)
    print(all_row_column)

def pdWay():
    file_dir = r"C:\Users\admin\Desktop\showdata.xlsx"
    df = pd.read_excel(file_dir, sheet_name='APIcase')

    data = df.head(3)
    # print('{}'.format(data))

    #1读取指定的单行,存列表
    row = df.loc[0].values
    # print(row)

    #2：读取指定的多行，数据会存在嵌套的列表里面
    more_row = df.loc[[0,1]].values
    # print(more_row)

    #3读取指定的行和列
    row_column = df.iloc[1, 2]
    # print(row_column)

    #4读取指定的多行多列值
    more_row_column = df.loc[[0, 1], ['Host', 'Method']].values
    # print(more_row_column)

    # 4获取所有行的指定列,没法获取字典格式
    all_row = df.loc[:,['Host', 'verify result ID']]
    # print(all_row)

    # 获取所有行号并打印
    row_id = df.index.values
    # print(row_id)

    # 获取列名并打印
    column_name = df.columns.values
    print(column_name)

    # 获取指定行数的值
    row_num = df.sample(2).values
    # print(row_num)

    # 获取指定列的值
    column_value = df['Host'].values
    # print(column_value)

    '''把Excel数据处理成字典'''
    test_dict = []
    for i in df.index.values:
        row_data = df.loc[i,['number', 'APIname', 'Host', 'RequestUrl', 'Method','RequestDataType',\
                             'RequestData', 'verifyResultID', 'verifyResultName']].to_dict()
        test_dict.append(row_data)
    print(test_dict)

def xlrdWay():

    file_dir = r"C:\Users\admin\Desktop\showdata.xlsx"
    data = xlrd.open_workbook(file_dir)

    table = data.sheet_by_name('APIcase')
    table = data.sheet_by_index(0)

    names = data.sheet_names()      #返回所有工作表的名字
    # print(names)

    rows = table.nrows      #获取有效行
    # print(rows)

    row_col = table.row(1)      ##返回由该行中所有的单元格对象组成的列表
    # print(row_col)

    p = table.row_values(1, start_colx=0, end_colx=None)     #获取第1行列的数据
    # print(p)

    L = table.row_len(2)        #获得第二列的长度
    # print(L)

    '''===================================================================='''

    col = table.ncols   #获取有效列
    # print(col)

    Y = table.col(2)        #返回该列所有数据
    # print(Y)

    u = table.col_values(7, start_rowx=0, end_rowx=None)     #获得第7列的数据
    # print(u)

    '''-----------------------------------------------------------------------'''
    t = table.cell(1, 9)        #返回单元格对象
    # print(t)

    k = table.cell_type(1, 6)       #返回单元格数据类型
    # print(k)

    v = table.cell_value(1, 6)      #返回单元格数据
    # print(v)


xlrdWay()